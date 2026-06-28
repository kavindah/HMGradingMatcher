#!/usr/bin/env python3
"""
grade_verify.py - verify a spec's grading against an approved reference spec
in the same grade-rule category.

CLI:    python grade_verify.py reference.pdf spec.pdf output.xlsx
Web:    verify_to_bytes(ref_fileobj, spec_fileobj, tol) -> (xlsx_bytes, summary)

For every measurement code the two PDFs share, the size-to-size grade
increments (deltas) are compared. Matches show green, mismatches red with the
expected value noted. A second sheet lists every exception. Increments are
compared (not absolute measurements), so two different garments in the same
grade-rule template verify correctly.
"""
import io, sys, re
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

SIZE_RE = re.compile(r"^(2XS|XS|S|M|L|XL|2XL|3XL|4XL|5XL)(?:-(S|T|TT))?$")
CODE_RE = re.compile(r"^[A-Z][A-Z0-9]*(_[A-Z0-9]+)*$")
STD_RUN = ["XXS","2XS","XS","S","M","L","XL","2XL","3XL","4XL"]


def norm(t): return re.sub(r"\s+"," ",(t or "")).strip()


def classify(header):
    roles,sizes={},{}
    for i,c in enumerate(header):
        c=norm(c); cl=c.lower()
        if not c: continue
        if cl.startswith("meas. code") or cl.startswith("meas.code") or cl=="code": roles["code"]=i
        elif "point of measurement" in cl or cl=="pom": roles["name"]=i
        elif SIZE_RE.match(c): sizes[i]=c
    return roles,sizes


def parse_banner(cells):
    j=" ".join(norm(c) for c in cells if c); info={}
    for k,p in (("set",r"Measurement Set:\s*([\w]+\s*:\s*[\w\- ]+?)\s+Sample"),
                ("sample",r"Sample Size:\s*([\w/]+)"),
                ("grade",r"Grade Rule Template:\s*(.+)$")):
        m=re.search(p,j)
        if m: info[k]=norm(m.group(1))
    return info


def extract(source):
    """source may be a file path OR a file-like object (BytesIO)."""
    sets={}
    def gs(n): return sets.setdefault(n,{"order":[],"rec":{},"meta":{}})
    with pdfplumber.open(source) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                roles,sizes,cur={},{},None
                for row in table:
                    if not row: continue
                    f=norm(row[0]).lower()
                    if f.startswith("measurement set:"):
                        info=parse_banner(row); cur=info.get("set",cur)
                        if cur:
                            meta=gs(cur)["meta"]
                            for k in ("sample","grade"):
                                if k in info and k not in meta: meta[k]=info[k]
                        continue
                    if f.startswith("meas. code") or f.startswith("meas.code") or f=="code":
                        roles,sizes=classify(row); continue
                    if "code" not in roles or not sizes: continue
                    code=re.sub(r"[\s*]+","",row[roles["code"]] or "")
                    if not CODE_RE.match(code): continue
                    if cur is None: cur="001 : Measurement"
                    b=gs(cur); name=norm(row[roles["name"]]) if "name" in roles else ""
                    nk=re.sub(r"[^a-z0-9]","",name.lower()); key=(code,nk)
                    if key not in b["rec"]:
                        b["order"].append(key); b["rec"][key]={"name":name,"vals":{}}
                    rec=b["rec"][key]
                    for ci,lab in sizes.items():
                        if ci<len(row):
                            v=norm(row[ci])
                            if v: rec["vals"].setdefault(lab,v)
    return sets


def num(v):
    try: return float(v)
    except: return None


def main_set(sets):
    best=None
    for n,b in sets.items():
        if "logo" in n.lower(): continue
        if best is None or len(b["order"])>len(sets[best]["order"]): best=n
    if best is None:  # only logos / nothing
        best=next(iter(sets)) if sets else None
    return (best, sets[best]) if best else (None,{"order":[],"rec":{},"meta":{}})


def deltas(rec, sizes):
    out={}
    for a,b in zip(sizes, sizes[1:]):
        va,vb=num(rec["vals"].get(a)), num(rec["vals"].get(b))
        if va is not None and vb is not None:
            out[f"{a}\u2192{b}"]=round(vb-va,2)
    return out


# ---------- styling ----------
FONT="Arial"
HDR=PatternFill("solid",fgColor="1F3864"); HDRF=Font(name=FONT,color="FFFFFF",bold=True,size=10)
GREEN=PatternFill("solid",fgColor="C6EFCE"); GREENF=Font(name=FONT,color="006100",size=10)
RED=PatternFill("solid",fgColor="FFC7CE");   REDF=Font(name=FONT,color="9C0006",bold=True,size=10)
GRAY=PatternFill("solid",fgColor="F2F2F2");  CELL=Font(name=FONT,size=10)
CODEF=Font(name=FONT,size=10,bold=True)
THIN=Side(style="thin",color="BFBFBF"); BORDER=Border(THIN,THIN,THIN,THIN)
CENTER=Alignment(horizontal="center",vertical="center")
LEFT=Alignment(horizontal="left",vertical="center")


def verify(ref_sets, spec_sets, tol=0.0):
    _,ref=main_set(ref_sets); _,spec=main_set(spec_sets)
    sizes=[s for s in STD_RUN if any(s in r["vals"] for r in spec["rec"].values())]
    breaks=[f"{a}\u2192{b}" for a,b in zip(sizes,sizes[1:])]
    rows=[]
    for key in spec["order"]:
        if key not in ref["rec"]: continue
        rec=spec["rec"][key]
        sd=deltas(rec,sizes); rd=deltas(ref["rec"][key],sizes)
        cells=[]
        for br in breaks:
            s=sd.get(br); r=rd.get(br)
            if s is None or r is None: cells.append((br,s,r,None))
            else: cells.append((br,s,r,abs(s-r)<=tol))
        rows.append({"code":key[0],"name":rec["name"],"cells":cells})
    only_spec=[(k[0],spec["rec"][k]["name"]) for k in spec["order"] if k not in ref["rec"]]
    only_ref =[(k[0],ref["rec"][k]["name"])  for k in ref["order"]  if k not in spec["rec"]]
    return breaks, rows, only_spec, only_ref


def build_report_wb(breaks, rows, only_spec, only_ref, ref_meta, spec_meta):
    wb=Workbook(); ws=wb.active; ws.title="Grade Check"
    ws.cell(row=1,column=1,value="Grade verification \u2014 spec increments vs approved reference").font=Font(name=FONT,bold=True,size=12)
    ws.cell(row=2,column=1,value=(f"Reference grade rule: {ref_meta.get('grade','?')}   |   "
        f"Spec grade rule: {spec_meta.get('grade','?')}")).font=Font(name=FONT,size=9,italic=True)
    if norm(ref_meta.get('grade',''))!=norm(spec_meta.get('grade','')):
        ws.cell(row=3,column=1,value="\u26a0 Grade rule templates differ \u2014 these two styles may not be the same category.").font=Font(name=FONT,size=9,color="9C0006",bold=True)
    hrow=5
    heads=["Meas. Code","Point of Measurement"]+breaks+["Status"]
    for c,h in enumerate(heads,1):
        cell=ws.cell(row=hrow,column=c,value=h); cell.fill=HDR; cell.font=HDRF
        cell.alignment=CENTER; cell.border=BORDER
    r=hrow+1
    for row in rows:
        ws.cell(row=r,column=1,value=row["code"]).font=CODEF
        nm=ws.cell(row=r,column=2,value=row["name"]); nm.font=CELL; nm.alignment=LEFT
        miss=0; comparable=0
        for j,(br,s,rv,ok) in enumerate(row["cells"]):
            cell=ws.cell(row=r,column=3+j)
            if ok is None:
                cell.value="" if s is None else s; cell.fill=GRAY; cell.font=CELL
            elif ok:
                cell.value=s; cell.fill=GREEN; cell.font=GREENF; comparable+=1
            else:
                cell.value=s; cell.fill=RED; cell.font=REDF; comparable+=1; miss+=1
                cell.comment=Comment(f"expected {rv} (reference), got {s}","grade_verify")
            cell.alignment=CENTER; cell.border=BORDER
        st=ws.cell(row=r,column=3+len(breaks))
        if comparable==0: st.value="\u2014"; st.font=CELL
        elif miss==0: st.value="\u2713 matches"; st.fill=GREEN; st.font=GREENF
        else: st.value=f"\u2717 {miss} mismatch"; st.fill=RED; st.font=REDF
        st.alignment=CENTER; st.border=BORDER
        for c in range(1,3+len(breaks)+1): ws.cell(row=r,column=c).border=BORDER
        r+=1
    ws.column_dimensions["A"].width=11; ws.column_dimensions["B"].width=34
    for j in range(len(breaks)): ws.column_dimensions[get_column_letter(3+j)].width=8
    ws.column_dimensions[get_column_letter(3+len(breaks))].width=13
    ws.freeze_panes=ws.cell(row=hrow+1,column=3)

    ex=wb.create_sheet("Exceptions")
    for c,h in enumerate(["Meas. Code","Point of Measurement","Size break","Reference \u0394","Spec \u0394","Difference"],1):
        cell=ex.cell(row=1,column=c,value=h); cell.fill=HDR; cell.font=HDRF; cell.alignment=CENTER; cell.border=BORDER
    er=2
    for row in rows:
        for (br,s,rv,ok) in row["cells"]:
            if ok is False:
                ex.cell(row=er,column=1,value=row["code"]).font=CODEF
                ex.cell(row=er,column=2,value=row["name"]).font=CELL
                ex.cell(row=er,column=3,value=br).font=CELL
                ex.cell(row=er,column=4,value=rv).font=CELL
                ex.cell(row=er,column=5,value=s).font=REDF
                ex.cell(row=er,column=6,value=round(s-rv,2)).font=CELL
                er+=1
    if er==2: ex.cell(row=2,column=1,value="No grading mismatches \u2014 every shared measurement matched the reference.").font=GREENF
    for col,w in (("A",11),("B",34),("C",12),("D",11),("E",10),("F",11)): ex.column_dimensions[col].width=w
    nr=er+2
    ex.cell(row=nr,column=1,value="Measurements only in SPEC (not in reference \u2014 not verified):").font=Font(name=FONT,bold=True,size=9)
    for k in only_spec: nr+=1; ex.cell(row=nr,column=1,value=k[0]).font=CELL; ex.cell(row=nr,column=2,value=k[1]).font=CELL
    nr+=2
    ex.cell(row=nr,column=1,value="Measurements only in REFERENCE (missing from spec):").font=Font(name=FONT,bold=True,size=9)
    for k in only_ref: nr+=1; ex.cell(row=nr,column=1,value=k[0]).font=CELL; ex.cell(row=nr,column=2,value=k[1]).font=CELL
    return wb


def _summary(breaks, rows, only_spec, only_ref, ref_meta, spec_meta):
    mm_breaks=sum(1 for row in rows for c in row["cells"] if c[3] is False)
    mm_rows=sum(1 for row in rows if any(c[3] is False for c in row["cells"]))
    return {
        "checked": len(rows),
        "mismatch_breaks": mm_breaks,
        "mismatch_rows": mm_rows,
        "only_spec": len(only_spec),
        "only_ref": len(only_ref),
        "ref_grade": ref_meta.get("grade",""),
        "spec_grade": spec_meta.get("grade",""),
        "grade_match": norm(ref_meta.get("grade",""))==norm(spec_meta.get("grade","")),
    }


def verify_to_bytes(ref_source, spec_source, tol=0.0):
    """In-memory: returns (xlsx_bytes, summary_dict). Nothing written to disk."""
    ref_sets=extract(ref_source); spec_sets=extract(spec_source)
    _,refm=main_set(ref_sets); _,specm=main_set(spec_sets)
    breaks,rows,os_,or_=verify(ref_sets,spec_sets,tol)
    if not rows and not os_:
        raise ValueError("No measurement tables found, or the two PDFs share no measurement codes.")
    wb=build_report_wb(breaks,rows,os_,or_,refm["meta"],specm["meta"])
    bio=io.BytesIO(); wb.save(bio)
    return bio.getvalue(), _summary(breaks,rows,os_,or_,refm["meta"],specm["meta"])


if __name__=="__main__":
    data,summary=verify_to_bytes(sys.argv[1], sys.argv[2])
    open(sys.argv[3],"wb").write(data)
    print(summary, "->", sys.argv[3])
